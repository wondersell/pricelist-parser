import csv
import xlrd
from abc import abstractmethod
from openpyxl import load_workbook

from .pricelist_item import PricelistParserItem

header_signatures = {
    'sku': ['код', 'артикул', 'модель', 'штрих-код'],
    'price': ['цена', 'ррц', 'стоимость', 'розница', 'рознич', 'оптовая'],
    'quantity': ['количество', 'кол-во', 'кол'],
    'name': ['название', 'наименование', 'товар', 'номенклатура', 'продукци'],
    'description': ['описание', 'характеристики'],
    'dimensions': ['размер', 'габариты'],
    'weight': ['вес'],
    'link': ['ссылка'],
    'vat': ['ндс'],
    'order': ['заказ', 'заявк', '№'],
}


class Extruder(object):
    required_headers = ['sku', 'price']

    def __init__(self, source_file=None, **kwargs):
        self.items = []
        self.loaded_file = None
        self.header_signatures = header_signatures.copy()
        self.source_file = source_file

    def update_header_signatures(self, signatures, replace=False):
        for key, signature in signatures.items():
            if key in self.header_signatures.keys() and not replace:
                self.header_signatures[key] += signature
            else:
                self.header_signatures[key] = signature

    def load_data(self, **kwargs):
        wb = self.load_file(self.source_file, **kwargs)

        # Загружаем данные о полях
        sheets_with_headers = self.get_headers_map(wb)

        if len(sheets_with_headers) == 0:
            raise ValueError('No worksheets with detectable data found')

        # Проходимся по всем листам с данными и загружаем из них товары
        for sheet_info in sheets_with_headers:
            self.items = self.items + self.extract_data_from_sheet(worksheet_info=sheet_info)

        return self

    def get_headers_map(self, workbook):
        sheet_names = self.get_worksheet_names(workbook)
        headers_map = []

        # Проверяем, на каких листах есть что-то похожее на данные о ценах
        # Для этого выбираем все листы, в которых мы можем распознать заголовки
        for sheet_name in sheet_names:
            ws = self.get_worksheet(wb=workbook, ws_name=sheet_name)
            headers = self.detect_headers(ws)

            if len(headers) > 0:
                headers = self.select_best_headers(headers)
                headers_map.append({
                    'worksheet': ws,
                    'headers': headers,
                })

        return headers_map

    def extract_data_from_sheet(self, worksheet_info):
        # Одна строка – один товар. Попутно валидируем данные, потому что поставщики любят вставлять пустые строки
        # для отбивки разных разделов
        extracted = []
        ws = worksheet_info['worksheet']
        for row_idx in range(worksheet_info['headers']['row'] + 1, self.get_rows_num(ws) + 1):
            extruded = {}

            for header_info in worksheet_info['headers']['headers']:
                col_idx = header_info['column']
                column_type = header_info['type']

                extruded[str(column_type)] = self.get_cell(ws=ws, col=col_idx, row=row_idx)

            pricelist_item = PricelistParserItem(**extruded)

            if pricelist_item.is_valid():
                extracted.append(pricelist_item)

        return extracted

    def detect_headers(self, worksheet):
        detected_header_rows = []

        # Для поиска проходимся по всем строкам таблицы. Строк с заголовками может найтись не одна, потом отсеем лишние
        for row_idx in range(1, self.get_rows_num(ws=worksheet) + 1):
            for col_idx in range(1, self.get_cols_num(worksheet) + 1):
                header_type = self.detect_header_type(str(self.get_cell(ws=worksheet, row=row_idx, col=col_idx)))

                # Если хотя бы одна ячейка в строке определилась как заголовок
                if header_type is not None:
                    # Записываем все непустые значения в качестве заголовков, попутно определяя ее тип,
                    # переходим к следующей строке
                    headers = []

                    for _col_idx in range(1, self.get_cols_num(worksheet) + 1):
                        cell_value = self.get_cell(worksheet, col=_col_idx, row=row_idx)

                        if cell_value and str(cell_value).strip() != '':
                            headers.append({
                                'column': _col_idx,
                                'name': cell_value,
                                'type': self.detect_header_type(cell_value),
                            })

                    detected_header_rows.append({
                        'row': row_idx,
                        'headers': headers,
                    })

                    break

        return detected_header_rows

    def select_best_headers(self, headers):
        # Если строка с заголовком только одна, то и вернем её
        if len(headers) == 1:
            return headers[0]

        # Вычислим количество найденных столбцов, тип которых мы смогли определить
        detected_headers = {}
        for variant in headers:
            detected = 0
            for header in variant['headers']:
                if header['type'] is not None:
                    detected += 1

            detected_headers[variant['row']] = detected

        # Отсортируем по количеству столбцов, чтобы найти строку с максимальным количеством заголовком
        detected_headers = {k: v for k, v in sorted(detected_headers.items(), key=lambda item: item[1])}
        row = detected_headers.popitem()[0]

        # Найдем и вернем эту строку
        for variant in headers:
            if variant['row'] == row:
                return variant

    def detect_header_type(self, cell_value):
        for header_type, headers in self.header_signatures.items():
            for substring in headers:
                if substring in str(cell_value).lower().strip():
                    return header_type

        return None

    @abstractmethod
    def load_file(self, filename, **kwargs):
        pass

    @abstractmethod
    def get_worksheet_names(self, wb):
        pass

    @abstractmethod
    def get_worksheet(self, wb, ws_name):
        pass

    @abstractmethod
    def get_cols_num(self, ws):
        pass

    @abstractmethod
    def get_rows_num(self, ws):
        pass

    @abstractmethod
    def get_cell(self, ws, col, row):
        pass


class XlsxExtruder(Extruder):
    def load_file(self, filename, **kwargs):
        self.loaded_file = load_workbook(filename=filename)
        return self.loaded_file

    def get_worksheet_names(self, wb):
        return wb.sheetnames

    def get_worksheet(self, wb, ws_name):
        return wb[ws_name]

    def get_cols_num(self, ws):
        return ws.max_column

    def get_rows_num(self, ws):
        return ws.max_row

    def get_cell(self, ws, col, row):
        return ws.cell(row=row, column=col).value


class XlsExtruder(Extruder):
    def load_file(self, filename, **kwargs):
        self.loaded_file = xlrd.open_workbook(filename)
        return self.loaded_file

    def get_worksheet_names(self, wb):
        return wb.sheet_names()

    def get_worksheet(self, wb, ws_name):
        return wb.sheet_by_name(ws_name)

    def get_cols_num(self, ws):
        return ws.ncols

    def get_rows_num(self, ws):
        return ws.nrows

    def get_cell(self, ws, col, row):
        return ws.cell(row - 1, col - 1).value


class CsvExtruder(Extruder):
    def __init__(self, source_file=None, **kwargs):
        self.data = []
        super().__init__(source_file, **kwargs)

    def load_file(self, filename, **kwargs):
        encoding = kwargs['encoding'] if 'encoding' in kwargs else 'utf-8'

        with open(filename, 'r', encoding=encoding) as file_object:
            reader = csv.reader(file_object)
            for row in reader:
                self.data.append(row)

        return self

    def get_worksheet_names(self, wb):
        return ['default']

    def get_worksheet(self, wb, ws_name='default'):
        return self

    def get_cols_num(self, ws):
        return len(ws.data[0])

    def get_rows_num(self, ws):
        return len(ws.data)

    def get_cell(self, ws, col, row):
        return ws.data[row - 1][col - 1]
